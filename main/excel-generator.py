import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import platform
import subprocess
import socket
import re
import os

def get_hostname():
    try:
        return socket.gethostname()
    except:
        return "Unknown"

def get_system_ip():
    try:
        # Get all network interfaces' IP addresses
        cmd = "ip addr show"
        result = subprocess.run(cmd.split(), capture_output=True, text=True)
        
        if result.returncode == 0:
            # Find IPv4 addresses that aren't 127.0.0.1 or 192.168.*
            pattern = r'inet\s+(\d+\.\d+\.\d+\.\d+)/'
            ips = re.findall(pattern, result.stdout)
            
            # Filter out localhost and 192.168.* addresses
            filtered_ips = [ip for ip in ips if not ip.startswith(('127.', '192.168.'))]
            
            if filtered_ips:
                return filtered_ips[0]
    except:
        pass
    return "Unknown"

def get_cpu_model():
    try:
        with open('/proc/cpuinfo', 'r') as f:
            for line in f:
                if line.startswith('model name'):
                    # Just remove the "model name      : " prefix
                    return line.split(':', 1)[1].strip()
    except Exception as e:
        print(f"Error reading CPU model: {str(e)}")
    return "Unknown"


def get_cpu_family():
    try:
        with open('/proc/cpuinfo', 'r') as f:
            for line in f:
                if line.startswith('cpu family'):
                    family = int(line.split(':')[1].strip())
                    return family, 'Genoa' if family == 25 else 'Turin' if family == 26 else 'Unknown'
    except Exception as e:
        print(f"Error reading CPU family: {str(e)}")
        return None, 'Unknown'

def get_socket_info():
    try:
        cmd = "lscpu"
        result = subprocess.run(cmd.split(), capture_output=True, text=True)
        
        sockets = "0"
        cores_per_socket = "0"
        
        if result.returncode == 0:
            for line in result.stdout.splitlines():
                if "Socket(s):" in line:
                    sockets = line.split(':')[1].strip()
                elif "Core(s) per socket:" in line:
                    cores_per_socket = line.split(':')[1].strip()
        
        return sockets, cores_per_socket
    except:
        return "0", "0"

def get_cpu_frequencies():
    try:
        base_path = "/sys/devices/system/cpu/cpu0/cpufreq"
        with open(f"{base_path}/scaling_min_freq", 'r') as f:
            min_freq = f.read().strip()
        with open(f"{base_path}/scaling_max_freq", 'r') as f:
            max_freq = f.read().strip()
        return f"{min_freq}", f"{max_freq}"
    except:
        return "Unknown", "Unknown"

def get_total_memory():
    try:
        with open('/proc/meminfo', 'r') as f:
            for line in f:
                if line.startswith('MemTotal'):
                    total_kb = int(line.split()[1])
                    return f"{total_kb/1024/1024:.2f}"
    except:
        return "Unknown"

def get_distro():
    try:
        # Try reading /etc/os-release first
        with open('/etc/os-release', 'r') as f:
            os_info = {}
            for line in f:
                if '=' in line:
                    key, value = line.rstrip().split('=', 1)
                    os_info[key] = value.strip('"')
            
            if 'PRETTY_NAME' in os_info:
                return os_info['PRETTY_NAME']
        
        # Fallback to lsb_release command
        result = subprocess.run(['lsb_release', '-d'], capture_output=True, text=True)
        if result.returncode == 0:
            return result.stdout.split(':')[1].strip()
        
        # Final fallback to platform module
        return f"{platform.linux_distribution()[0]} {platform.linux_distribution()[1]}"
    except:
        return "Linux"

def read_file(filename):
    with open(filename, 'r') as f:
        return [line.strip() for line in f if line.strip()]

def get_kernel_version(filename):
    try:
        with open(filename, 'r') as f:
            version = f.read().strip()
            # Extract version in x.x.x format
            parts = version.split('.')
            if len(parts) >= 3:
                return f"{parts[0]}.{parts[1]}.{parts[2].split('_')[0]}"
    except Exception as e:
        print(f"Error reading kernel version: {str(e)}")
        return "N/A"

def pad_array(arr, length, pad_value=''):
    return arr + [pad_value] * (length - len(arr))

def create_excel():
    try:
        # Get all system information
        hostname = get_hostname()
        system_ip = get_system_ip()
        cpu_model = get_cpu_model()
        family, family_name = get_cpu_family()
        sockets, cores_per_socket = get_socket_info()
        min_freq, max_freq = get_cpu_frequencies()
        total_memory = get_total_memory()
        
        # Get distribution info
        distro = get_distro()

        # Get kernel version
        kernel_version = get_kernel_version('/var/lib/lkp-automation-data/state-files/kernel-version')

        # Read all files
        test_suites = read_file('/var/lib/lkp-automation-data/results/test_suites')
        base_without_vms = read_file('/var/lib/lkp-automation-data/results/Base-without_vms')
        patch_without_vms = read_file('/var/lib/lkp-automation-data/results/Patch-without_vms')
        base_with_vms = read_file('/var/lib/lkp-automation-data/results/Base-with_vms')
        patch_with_vms = read_file('/var/lib/lkp-automation-data/results/Patch-with_vms')
        base_with_lkp_vms = read_file('/var/lib/lkp-automation-data/results/Base-with_lkp_vms')
        patch_with_lkp_vms = read_file('/var/lib/lkp-automation-data/results/Patch-with_lkp_vms')

        # Process test_suites into two columns
        test_suite_col = []
        test_col = []
        
        for line in test_suites[1:]:  # Skip header
            if line.startswith(','):
                test_suite_col.append('')
                test_col.append(line[1:])
            else:
                parts = line.split(',')
                if len(parts) > 1:
                    test_suite_col.append(parts[0])
                    test_col.append(parts[1])
                else:
                    test_suite_col.append(parts[0])
                    test_col.append('')

        # Find the maximum length needed
        max_length = max(
            len(test_suite_col),
            len(test_col),
            len(base_without_vms) - 1,
            len(patch_without_vms) - 1,
            len(base_with_vms) - 1,
            len(patch_with_vms) - 1,
            len(base_with_lkp_vms) - 1,
            len(patch_with_lkp_vms) - 1
        )

        # Pad all arrays
        test_suite_col = pad_array(test_suite_col, max_length)
        test_col = pad_array(test_col, max_length)
        
        # Create DataFrame with padded arrays
        df = pd.DataFrame({
            'Test_Suites': test_suite_col,
            'Test': test_col,
            'Base-without_vms': pad_array(base_without_vms[1:], max_length),
            'Patch-without_vms': pad_array(patch_without_vms[1:], max_length),
            'Base-with_vms': pad_array(base_with_vms[1:], max_length),
            'Patch-with_vms': pad_array(patch_with_vms[1:], max_length),
            'Base-with_lkp_vms': pad_array(base_with_lkp_vms[1:], max_length),
            'Patch-with_lkp_vms': pad_array(patch_with_lkp_vms[1:], max_length)
        })

        # Save to Excel directly in the current directory
        excel_file = '/var/lib/lkp-automation-data/results/raw-LKP-results.xlsx'
        df.to_excel(excel_file, index=False)

# Load the workbook to apply formatting
        wb = load_workbook(excel_file)
        ws = wb.active

        # Insert row at the very top for title
        ws.insert_rows(1)
        
        # Merge cells A1:H1
        ws.merge_cells('A1:H1')
        
        # Set the merged cell value and formatting
        merged_cell = ws['A1']
        merged_cell.value = f"LKP tests results for {family_name} system with {distro}"
        
        # Set font properties for title
        merged_cell.font = Font(size=20, bold=True)
        
        # Set alignment for title
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set row height for title
        ws.row_dimensions[1].height = 25.5
        
        # Set background color to bright yellow for title
        bright_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        merged_cell.fill = bright_yellow

        # Add border to the merged cell
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        merged_cell.border = border

        # Insert new row for system details
        ws.insert_rows(2)
        
        # Merge cells A2:H2 for system details
        ws.merge_cells('A2:H2')
        
        # Create system details content with different font sizes
        system_details = (
            f"{family_name} system details\n"  # First line (size 14)
            f"Host Name: {hostname} [ IP: {system_ip} ], Model name: {cpu_model}\n"  # Second line
            f"CPU family: {family}, Sockets/Core: {sockets} sockets/{cores_per_socket} cores per socket, "
            f"CPU Freq: Max - {max_freq} KHz Min - {min_freq} KHz\n"  # Third line
            f"Memory: {total_memory} GiB, Disks: NVMe SSD disks"  # Fourth line
        )
        
        # Set the system details cell
        system_details_cell = ws['A2']
        system_details_cell.value = system_details
        
        # Set font properties (size 10, bold)
        system_details_cell.font = Font(size=10, bold=True)
        
        # Set alignment with wrap text
        system_details_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Set row height to 3.5 cm (approximately 99 points)
        ws.row_dimensions[2].height = 65
        
        # Set background color to bright yellow
        system_details_cell.fill = bright_yellow
        
        # Add border to the system details cell
        system_details_cell.border = border

        # Insert new row for test information
        ws.insert_rows(3)
        # Set row height to accommodate 3 lines of text size 13
        ws.row_dimensions[3].height = 60  # Approximately 20 points per line for 3 lines

        # Format Column 1 (Test Suite)
        cell_a3 = ws['A3']
        cell_a3.value = "Test Suite"
        cell_a3.font = Font(size=22, bold=True)
        cell_a3.alignment = Alignment(horizontal='left', vertical='top')
        # Set column A width to exactly fit "Test Suite" in size 22 plus 0.5 cm
        ws.column_dimensions['A'].width = 17.3  # 15 (original width) + 2.3 (0.5 cm)

        # Format Column 2 (Tests)
        cell_b3 = ws['B3']
        cell_b3.value = "Tests"
        cell_b3.font = Font(size=22, bold=True)
        cell_b3.alignment = Alignment(horizontal='left', vertical='top')

# Merge and format columns 3-4
        ws.merge_cells('C3:D3')
        merged_cell_1 = ws['C3']
        merged_cell_1.value = f"Host with {distro} - no VMs"
        merged_cell_1.font = Font(size=13, bold=True)
        merged_cell_1.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

        # Merge and format columns 5-6
        ws.merge_cells('E3:F3')
        merged_cell_2 = ws['E3']
        merged_cell_2.value = f"Host with {distro} - with VMs\n[LKP run on Host only]"
        merged_cell_2.font = Font(size=13, bold=True)
        merged_cell_2.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

        # Merge and format columns 7-8
        ws.merge_cells('G3:H3')
        merged_cell_3 = ws['G3']
        merged_cell_3.value = f"Host with {distro} - with VMs\n[LKP run on Host + VMs]"
        merged_cell_3.font = Font(size=13, bold=True)
        merged_cell_3.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

        # Apply sea blue background to merged cells with black text
        sea_blue = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
        for cell in [merged_cell_1, merged_cell_2, merged_cell_3]:
            cell.fill = sea_blue
            cell.font = Font(size=13, bold=True, color='000000')  # Black text for better visibility

        # Delete the first row and insert a new one for kernel version
        ws.delete_rows(4)
        ws.insert_rows(4)

        # Add kernel version information
        kernel_headers = {
            3: f"kernel ver: {kernel_version} (Base-kernel)",
            4: f"kernel ver: {kernel_version} (kernel-with-patches)",
            5: f"kernel ver: {kernel_version} (Base-kernel)",
            6: f"kernel ver: {kernel_version} (kernel-with-patches)",
            7: f"kernel ver: {kernel_version} (Base-kernel)",
            8: f"kernel ver: {kernel_version} (kernel-with-patches)"
        }

        # Apply formatting to the kernel version row
        sea_blue = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
        bold_font = Font(bold=True)
        left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

        for col, text in kernel_headers.items():
            cell = ws.cell(row=4, column=col, value=text)
            cell.fill = sea_blue
            cell.font = bold_font
            cell.alignment = left_align

        # Insert new row for "Run Time(sec)" labels
        ws.insert_rows(5)
        
        # Add "Run Time(sec)" labels
        for col in range(3, 9):  # Columns C to H
            cell = ws.cell(row=5, column=col, value="Run Time(sec)")
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)

        # Define colors
        light_orange = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
        light_blue = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        light_violet = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        very_light_green = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')

        # Format cell A11 (adjusted for new rows at top)
        cell_a11 = ws['A14']
        cell_a11.fill = light_violet
        cell_a11.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        cell_a11.font = Font(bold=True, size=22)

        # Insert empty rows before and after row 11 (adjusted for new rows at top)
        ws.insert_rows(14)
        ws.insert_rows(16)

        # Merge cells and apply formatting for rows 3-10 (adjusted for new rows at top)
        ws.merge_cells('A6:A13')
        merged_cell = ws['A6']
        merged_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        merged_cell.font = Font(bold=True, size=22)
        merged_cell.fill = light_orange

        # Set cell A13 to horizontal alignment (adjusted for new rows at top)
        cell_a13 = ws['A15']
        if cell_a13:
            cell_a13.alignment = Alignment(horizontal='center', vertical='center', text_rotation=0)
            cell_a13.font = Font(bold=True, size=22)

        # Merge cells and apply formatting for rows 14-41 (adjusted for new rows at top)
        ws.merge_cells('A17:A44')
        merged_cell = ws['A17']
        merged_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        merged_cell.font = Font(bold=True, size=22)
        merged_cell.fill = light_blue

        # Set column widths
        for col in range(3, 9):  # Columns C to H
            ws.column_dimensions[get_column_letter(col)].width = 15

        ws.column_dimensions['B'].width = 23  # Width to fit "100%-300s-whetstone-double"

        # Convert numeric values and apply left alignment
        left_align = Alignment(horizontal='left', vertical='center')
        for row in ws.iter_rows(min_row=6, max_row=44, min_col=3, max_col=8):
            for cell in row:
                cell.alignment = left_align
                if cell.value and isinstance(cell.value, str) and cell.value.strip():
                    try:
                        cell.value = float(cell.value)
                    except ValueError:
                        pass

        # Color rows with very light green background
        for row_num in [14, 16]:
            for cell in ws[row_num]:
                cell.fill = very_light_green

        # Set vertical text rotation for column A
        vertical_alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        for row in ws.iter_rows(min_row=6, max_row=44, min_col=1, max_col=1):
            for cell in row:
                if cell.value and cell.row != 15:  # Skip row 15 (adjusted for new rows at top)
                    cell.alignment = vertical_alignment

        # Apply borders to all cells in the data range
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=1, max_row=44, min_col=1, max_col=8):
            for cell in row:
                cell.border = border

        # Save the formatted workbook
        wb.save(excel_file)
        print(f"Excel file '{excel_file}' has been created successfully!")

    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    create_excel()
