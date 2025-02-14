from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def calculate_percentage(numerator_val, denominator_val):
    try:
        if denominator_val == 0:
            return None
        result = (numerator_val - denominator_val) / denominator_val
        return result
    except:
        return None

def modify_excel():
    file_path = "/var/lib/lkp-automation-data/results/raw-LKP-results.xlsx"
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Part 1: Header and row 5 formatting
    ws.merge_cells('K3:M4')
    merged_cell = ws['K3']
    merged_cell.value = "Run to Run variance"
    merged_cell.font = Font(bold=True, size=12)
    merged_cell.fill = PatternFill(start_color='87CEEB', 
                                 end_color='87CEEB',
                                 fill_type='solid')
    merged_cell.alignment = Alignment(horizontal='center', 
                                    vertical='center')
    
    # Set column widths
    column_width = 19
    for col in ['K', 'L', 'M']:
        ws.column_dimensions[col].width = column_width
    
    # Row 5 headers
    cell_values = {
        'K5': "No VMs",
        'L5': "LKP run on host only",
        'M5': "LKP run on host + VMs"
    }
    
    # Format row 5 cells
    light_blue = 'B0E2FF'
    for cell_pos, value in cell_values.items():
        cell = ws[cell_pos]
        cell.value = value
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=light_blue,
                               end_color=light_blue,
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='center',
                                 vertical='center',
                                 wrap_text=True)
    
    # Part 2: Calculations and formatting
    red_color = 'FF0000'
    light_green = '90EE90'
    row_ranges = [(6, 13), (15, 15), (17, 44)]
    
    column_mappings = {
        'K': ('D', 'C'),
        'L': ('F', 'E'),
        'M': ('H', 'G')
    }
    
    # Process each column with its specific calculation
    for target_col, (num_col, den_col) in column_mappings.items():
        for start_row, end_row in row_ranges:
            for row in range(start_row, end_row + 1):
                numerator_cell = ws[f'{num_col}{row}']
                denominator_cell = ws[f'{den_col}{row}']
                
                try:
                    numerator_val = float(numerator_cell.value) if numerator_cell.value is not None else None
                    denominator_val = float(denominator_cell.value) if denominator_cell.value is not None else None
                    
                    if numerator_val is not None and denominator_val is not None:
                        result = calculate_percentage(numerator_val, denominator_val)
                        target_cell = ws[f'{target_col}{row}']
                        
                        if result is not None:
                            target_cell.value = result
                            target_cell.number_format = '0.00%'
                            
                            if result > 0.1 or result < -0.1:
                                background_color = red_color
                            else:
                                background_color = light_green
                                
                            target_cell.fill = PatternFill(start_color=background_color,
                                                         end_color=background_color,
                                                         fill_type='solid')
                        else:
                            target_cell.value = "N/A"
                except (ValueError, TypeError):
                    ws[f'{target_col}{row}'].value = "N/A"
    
    # Add borders to all cells in range K3:M44
    for col in ['K', 'L', 'M']:
        for row in range(3, 45):  # 3 to 44 inclusive
            cell = ws[f'{col}{row}']
            cell.border = thin_border
    
    # Save the modified file
    wb.save('/var/lib/lkp-automation-data/results/LKP-results.xlsx')
    print("File has been modified and saved as 'LKP-results.xlsx'")

# Run the function
modify_excel()
